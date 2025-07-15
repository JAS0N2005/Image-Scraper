# image_scraper/processor.py

import os
from openpyxl import Workbook, load_workbook
from config_manager import update_last_row
from processor_core import RowProcessor

def append_to_excel(path, entries):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.append(['row','activity_id','url','file','status','error'])
    else:
        wb = load_workbook(path)
        ws = wb.active
    for e in entries:
        ws.append([e['row'], e['activity_id'], e['url'],
                   e['file'], e['status'], e['error']])
    wb.save(path)

async def process_row(session, idx, type_, activity_id, website, summary, cfg):
    # 1) prepare per-row log
    log_rows = []

    # 2) instantiate and run the core logic
    rp = RowProcessor(
        session=session,
        idx=idx,
        type_=type_,
        activity_id=activity_id,
        website=website,
        cfg=cfg,
        log_rows=log_rows
    )
    await rp.run()

    # 3) write out log & summary & advance pointer
    append_to_excel(cfg['log_excel'], log_rows)
    summary.append({'row': idx, 'successes': rp.success, 'failures': rp.failures})
    update_last_row(idx)
    print(f"Row {idx} (ID {activity_id}): {rp.success} succeeded, {rp.failures} failed")
